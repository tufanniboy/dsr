"""
Export service — generates Excel, PDF, and CSV reports.
"""

import io
import csv
from typing import List, Dict, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """Handles exporting DSR data to Excel, PDF, and CSV formats."""

    def export_excel(self, data: List[Dict], title: str = "DSR Report") -> bytes:
        """Generate an Excel file from DSR data."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "DSR Report"

        # Styles
        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        currency_format = '₹#,##0.00'

        # Title row
        ws.merge_cells("A1:L1")
        title_cell = ws["A1"]
        title_cell.value = f"{title} — Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="FF6B00")
        title_cell.alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            "Date", "Manager", "Duty", "Start Reading", "End Reading",
            "Testing", "Rate", "Total Amount", "Card", "UPI",
            "Expenses", "Credit", "Cash In Hand", "Short Amount"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        row_num = 4
        for record in data:
            ws.cell(row=row_num, column=1, value=record.get("date", ""))
            ws.cell(row=row_num, column=2, value=record.get("manager_name", ""))
            ws.cell(row=row_num, column=3, value=record.get("duty_number", ""))
            ws.cell(row=row_num, column=4, value=record.get("start_reading", 0))
            ws.cell(row=row_num, column=5, value=record.get("end_reading", 0))
            ws.cell(row=row_num, column=6, value=record.get("testing", 0))
            ws.cell(row=row_num, column=7, value=record.get("rate", 0))

            for col_idx in [8, 9, 10, 11, 12, 13, 14]:
                cell = ws.cell(row=row_num, column=col_idx)
                cell.number_format = currency_format

            ws.cell(row=row_num, column=8, value=record.get("total_amount", 0))
            ws.cell(row=row_num, column=9, value=record.get("card", 0))
            ws.cell(row=row_num, column=10, value=record.get("upi", 0))
            ws.cell(row=row_num, column=11, value=record.get("expenses", 0))
            ws.cell(row=row_num, column=12, value=record.get("credit", 0))
            ws.cell(row=row_num, column=13, value=record.get("total_cash_in_hand", 0))
            ws.cell(row=row_num, column=14, value=record.get("short_amount", 0))

            for col in range(1, 15):
                ws.cell(row=row_num, column=col).border = thin_border

            row_num += 1

        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

        # Summary row
        if data:
            row_num += 1
            ws.cell(row=row_num, column=7, value="TOTAL:").font = Font(bold=True)
            for col, key in [(8, "total_amount"), (9, "card"), (10, "upi"),
                             (11, "expenses"), (12, "credit"), (13, "total_cash_in_hand")]:
                total = sum(float(r.get(key, 0) or 0) for r in data)
                cell = ws.cell(row=row_num, column=col, value=total)
                cell.font = Font(bold=True)
                cell.number_format = currency_format

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def export_csv(self, data: List[Dict]) -> str:
        """Generate CSV string from DSR data."""
        output = io.StringIO()
        if not data:
            return ""

        fieldnames = [
            "date", "manager_name", "duty_number", "start_reading", "end_reading",
            "testing", "rate", "total_amount", "card", "upi",
            "expenses", "credit", "total_cash_in_hand", "short_amount"
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)

        return output.getvalue()

    def export_pdf(self, data: List[Dict], title: str = "DSR Report") -> bytes:
        """Generate a PDF report from DSR data."""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=15 * mm)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Heading1"],
            fontSize=16, textColor=colors.HexColor("#FF6B00"),
            spaceAfter=10 * mm
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}",
            styles["Normal"]
        ))
        elements.append(Spacer(1, 10 * mm))

        # Table
        headers = [
            "Date", "Manager", "Duty", "Start", "End",
            "Testing", "Rate", "Total", "Card", "UPI",
            "Expenses", "Credit", "Cash", "Short"
        ]

        table_data = [headers]
        for record in data:
            table_data.append([
                record.get("date", ""),
                record.get("manager_name", "")[:15],
                record.get("duty_number", ""),
                record.get("start_reading", ""),
                record.get("end_reading", ""),
                record.get("testing", ""),
                record.get("rate", ""),
                f"₹{float(record.get('total_amount', 0) or 0):,.0f}",
                f"₹{float(record.get('card', 0) or 0):,.0f}",
                f"₹{float(record.get('upi', 0) or 0):,.0f}",
                f"₹{float(record.get('expenses', 0) or 0):,.0f}",
                f"₹{float(record.get('credit', 0) or 0):,.0f}",
                f"₹{float(record.get('total_cash_in_hand', 0) or 0):,.0f}",
                f"₹{float(record.get('short_amount', 0) or 0):,.0f}",
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF6B00")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF7ED")]),
        ]))

        elements.append(table)
        doc.build(elements)
        return buffer.getvalue()
